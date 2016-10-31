import datetime
import os,sys
import csv
from openpyxl import load_workbook

settings = sys.argv[1]
xls_file = sys.argv[2]

print("loading django",settings)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", settings)
import django
django.setup()


from aristotle_mdr.contrib.slots.models import Slot, SlotDefinition
from aristotle_mdr import models 
from aristotle_mdr.contrib.identifiers import models as MDR_ID
from comet import models as comet
from mallard_qr import models as mallard_qr
from logicaloutcomes import models as lo_models

lo,c = models.RegistrationAuthority.objects.get_or_create(name="Logical Outcomes")
lo_org = models.Organization.objects.get(pk=lo.pk)
lo_namespace, c = MDR_ID.Namespace.objects.get_or_create(naming_authority=lo_org, shorthand_prefix='lo')
# wg,c = models.Workgroup.objects.get_or_create(name="Import Workgroup")
# wg.registrationAuthorities=[lo]
# user = models.RegistrationAuthority.objects.get_or_create(name="Logical Outcomes")

default_definition = "Imported from the Indicator Reference Sheet"

def get_from_identifier(ident):
    obj = MDR_ID.ScopedIdentifier.objects.filter(
        namespace=lo_namespace,
        identifier=ident
    ).first()
    if obj is None:
        return None
    else:
        return obj.concept.item

def make_identifier(ident, item):
    obj = MDR_ID.ScopedIdentifier.objects.create(
        namespace=lo_namespace,
        identifier=ident,
        concept=item
    )
    return obj

def process_value_domains(wb):
    sheet = wb.get_sheet_by_name('Value Domain')

    val_dom = None
    n = 0
    for i,row in enumerate(sheet.iter_rows(row_offset=2)):
        meaning = row[1].value
        value = row[2].value or row[3].value or "-99"
        order = row[3].value or n
        if val_dom is None:
            vd_name = row[0].value
            dt,c = models.DataType.objects.get_or_create(name='Number')
            if c:
                register(dt)
            val_dom, created = models.ValueDomain.objects.get_or_create(
                    name = vd_name,
                    defaults = {
                        # "workgroup": wg,
                        "definition": default_definition,
                        "data_type": dt,
                    }
                )
            if not created:
                val_dom.permissiblevalue_set.all().delete()
            register(val_dom)
            models.PermissibleValue.objects.create(valueDomain=val_dom,value=value,meaning=meaning,order=order)
            n += 1
        elif row[0].value is None and row[1].value is None:
            if n == 1:
                val_dom.permissiblevalue_set.all().delete()
            val_dom = None
            n = 0
        else:
            models.PermissibleValue.objects.create(valueDomain=val_dom,value=value,meaning=meaning,order=order)
            n += 1

def process_data_elements(wb):
    sheet = wb.get_sheet_by_name('Data Elements')
    for row in sheet.iter_rows(row_offset=1):
        if row[1].value is None:
            continue
        name = get_col(row,'A').value #[1].value
        vd_name = get_col(row,'E')
        if vd_name:
            vd = models.ValueDomain.objects.get(name=vd_name.value.strip())
        else:
            vd = None

        defaults = {
            "name": name,
            "valueDomain": vd, #row["Short name:"],
            "definition": default_definition,
        }
        de_identifier = get_col(row,'B').value
        de = get_from_identifier(de_identifier)
        if de:
            models.DataElement.objects.filter(pk=de.pk).update(**defaults)
            de = comet.Indicator.objects.get(pk=de.pk)
        else:
            de = models.DataElement(**defaults)
            de.save()
            register(de)
            make_identifier(de_identifier,de)

        register(de)
        q_name = get_col(row,'D')
        if q_name:
            q, c = mallard_qr.Question.objects.update_or_create(
                name=get_col(row,'D').value.split('?')[0],
                question_text=get_col(row,'D').value,
                defaults = {
                    "collected_data_element":de,
                },
            )
            q.save()
            register(q)

        if vd and q_name:
            rd, c = mallard_qr.ResponseDomain.objects.update_or_create(
                question=q,
                value_domain=vd,
            )

        ind = get_from_identifier(get_col(row,'C').value)
        if ind:
            ind.numerators.add(de)
            ind.save()


def process_indicators(wb):
    sheet = wb.get_sheet_by_name('Indicators')

    slot_type_toc,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="Theory of Change",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )
    slot_type_np,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="No Poverty",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )
    slot_type_question,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="Question text",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )
    slot_type_dcm,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="Data collection method",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )
    slot_type_mom,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="Method of Measurement",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )
    slot_type_tou,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="Terms of use",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )
    slot_type_lang,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="Languages",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )
    slot_type_pop,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="Population",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )
    slot_type_rat,created = SlotDefinition.objects.get_or_create(
        app_label='comet',
        concept_type='indicator',
        slot_name="Rationale",
        cardinality = SlotDefinition.CARDINALITY.repeatable
    )

    for row in sheet.iter_rows(row_offset=1):
        if row[0].value is None:
            continue
        name = get_col(row,'A').value
        print name
        if not name:
            continue

        defn = get_col(row,'D')
        if defn:
            defn = defn.value or default_definition
        else:
            defn = default_definition
        defaults = {
            "name": name,
            "definition": defn,
            "references": get_col(row,'K').value,
            "rationale": get_col(row,'P').value if get_col(row,'M') else "",
            #"short_name": get_col(row,'C').value if get_col(row,'C') else "",
            "comments": get_col(row,'J').value if get_col(row,'C') else "",
        #     "disaggregation_description": lb_2_p(row["Disaggregation(s) 1:"])+lb_2_p(row["Disaggregation(s) 2:"]),
            "denominator_description": get_col(row,'F').value if get_col(row,'F') else "",
            "numerator_description": get_col(row,'E').value if get_col(row,'E') else "",
        }
        ind_identifier = get_col(row,'B').value
        ind = get_from_identifier(ind_identifier)
        if ind:
            comet.Indicator.objects.filter(pk=ind.pk).update(**defaults)
            ind = comet.Indicator.objects.get(pk=ind.pk)
        else:
            ind = comet.Indicator(**defaults)
            ind.save()
            register(ind)
            make_identifier(ind_identifier,ind)

        des = [de.strip() for de in get_col(row,'F').value.split(';')]
        for de in des:
            de = models.DataElement.objects.filter(short_name=de).first()
            if de:
                print "  -",de
                ind.numerators.add(de)
        
        instument_name = get_col(row,'I').value
        instrument = lo_models.Instrument.objects.filter(name=instument_name)
        if instrument.exists():
            instrument.first().indicators.add(ind)
        else:
            print "No instrument", instument_name, instrument

        text_to_slots(ind, get_col(row,'Q').value, slot_type_np)
        text_to_slots(ind, get_col(row,'O').value, slot_type_toc)
        text_to_slots(ind, get_col(row,'H').value, slot_type_dcm)
        text_to_slots(ind, get_col(row,'C').value, slot_type_question,clean=True)
        text_to_slots(ind, get_col(row,'G').value, slot_type_mom,clean=True)
        text_to_slots(ind, get_col(row,'N').value, slot_type_tou)
        text_to_slots(ind, get_col(row,'J').value, slot_type_lang)
        text_to_slots(ind, get_col(row,'L').value, slot_type_pop)
        text_to_slots(ind, get_col(row,'M').value, slot_type_rat)
        
        if get_col(row,'P'):
            goal_name = get_col(row,'P').value
            goal = lo_models.Goal.objects.get(short_name=goal_name)
            goal.indicators.add(ind)
            goal.save()

def text_to_slots(item, col,slot_type,clean=False):
    print slot_type, col
    for val in col.split(';'):
        val = val.strip()
        if clean:
            val = lb_2_p(val, sep="\n")
        if val:
            obj,created = Slot.objects.get_or_create(
                type=slot_type,
                concept=item,
                value=val
            )

def process_instruments(wb):
    sheet = wb.get_sheet_by_name('Instruments')
    for row in sheet.iter_rows(row_offset=1):
        if row[2].value is None:
            continue
        name = get_col(row,'A').value

        defaults = {
            "definition": lb_2_p(get_col(row,'E').value or default_definition, sep="\n"),
            "terms_of_use": get_col(row,'D').value or "",
            "where_to_get": get_col(row,'C').value or "",
            "references": get_col(row,'B').value or "",
        }
        inst, c = lo_models.Instrument.objects.update_or_create(
            name=name,
            **defaults
        )
        register(inst)


def process_goals(wb):
    sheet = wb.get_sheet_by_name('Sustainable Development Goals')
    for row in sheet.iter_rows(row_offset=1):
        if row[0].value is None:
            continue

        goal, c = lo_models.Goal.objects.get_or_create(
            short_name = get_col(row,'A').value,
            name = get_col(row,'B').value,
            definition = get_col(row,'C').value,
            origin_URI = get_col(row,'D').value,
        )
        register(goal)

def register(thing):
    models.Status.objects.get_or_create(
        concept=thing,
        registrationAuthority=lo,
        registrationDate=datetime.date(2009, 4, 28),
        state=models.STATES.recorded
    )



def lb_2_p(txt, sep="\n\n"):
    if sep in txt:
        return "<p>"+"</p><p>".join([l for l in txt.split(sep) if l != ""])+"</p>"
    else:
        return txt

import re
def jpger(txt):
    txt = re.sub(r"(https?://.*?jpg)", "<img src='\\1'>", txt)
    return txt

def get_col(row,col):
    for cell in row:
        if cell.value is not None:
            if cell.coordinate.startswith(col):
                return cell


if __name__ == "__main__":

    print("starting")
    wb = load_workbook(xls_file, read_only=True)

    process_goals(wb)
    process_instruments(wb)
    process_value_domains(wb)
    process_indicators(wb)
    process_data_elements(wb)
    
