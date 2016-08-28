import datetime
import os,sys
import csv
from openpyxl import load_workbook

settings = sys.argv[1]
xls_file = sys.argv[2]

print("loading django",settings)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", settings)
import django
django.setup()


from aristotle_mdr.contrib.slots.models import Slot, SlotDefinition
from aristotle_mdr import models 
from comet import models as comet
from mallard_qr import models as mallard_qr
from logicaloutcomes import models as lo_models

lo,c = models.RegistrationAuthority.objects.get_or_create(name="Logical Outcomes")
wg,c = models.Workgroup.objects.get_or_create(name="Import Workgroup")
wg.registrationAuthorities=[lo]
user = models.RegistrationAuthority.objects.get_or_create(name="Logical Outcomes")

default_definition = "Imported from the Indicator Reference Sheet"

slots = dict()
slot_columns = [
    ('AF','Terms of use',''),
    # (u'Indicator Code:','Code',''),
    # (u'Citation instruction:','Citation instruction','How to cite this document'),
    # (u'Usage instruction:','Usage instruction','How to use this document'),
    # (u'Method of Measurement 1:','Method of Measurement',''),
    # (u'Method of Measurement 2:','Method of Measurement',''),
    # (u'Data Collection Frequency 1:','Data Collection Frequency',''),
    # (u'Data Collection Frequency 2:','Data Collection Frequency',''),
    # (u'Data Collection Frequency 3:','Data Collection Frequency',''),
    # (u'Data Source:','Data Source',''),
    # ('Instrument:','Instrument',''),
    # ('DOI:','DOI',''),
    # ('Universe:','Universe','')
]

def process_value_domains(wb):
    
    sheet = wb.get_sheet_by_name('Answer Options')
    #for row in sheet.iter_rows(row_offset=3):
    #    if row[1].value is None:
    #        continue    #vds = csvfile.read().split('\n\t\t\t\n')
    val_dom = None
    n = 0
    for i,row in enumerate(sheet.iter_rows(row_offset=2)):
        meaning = row[1].value
        value = row[2].value or row[3].value or "-99"
        order = row[3].value or n
        if val_dom is None:
            vd_name = row[0].value
            dt,c = models.DataType.objects.get_or_create(name='Number')
            if c:
                register(dt)
            val_dom, created = models.ValueDomain.objects.get_or_create(
                    name = vd_name,
                    defaults = {
                        # "workgroup": wg,
                        "definition": default_definition,
                        "data_type": dt,
                    }
                )
            if not created:
                val_dom.permissiblevalue_set.all().delete()
            register(val_dom)
            models.PermissibleValue.objects.create(valueDomain=val_dom,value=value,meaning=meaning,order=order)
            n += 1
        elif row[0].value is None and row[1].value is None:
            if n == 1:
                val_dom.permissiblevalue_set.all().delete()
            val_dom = None
            n = 0
        else:
            models.PermissibleValue.objects.create(valueDomain=val_dom,value=value,meaning=meaning,order=order)
            n += 1

def process_data_elements(wb):
    sheet = wb.get_sheet_by_name('Data Elements')
    for row in sheet.iter_rows(row_offset=3):
        if row[1].value is None:
            continue
        name = get_col(row,'B').value #[1].value
        vd_name = get_col(row,'O')
        if vd_name:
            vd = models.ValueDomain.objects.get(name=vd_name.value.strip())
        else:
            vd = None

        defaults = {
            "valueDomain": vd, #row["Short name:"],
            "definition": default_definition,
            "short_name": get_col(row,'D').value,
        #     "definition": row["Purpose:"],
        #     "references": row["References:"],
        #     "rationale": lb_2_p(row["Interpretation:"]),
        #     "numerator_description": lb_2_p(row["Numerator:"]),
        #     "denominator_description": lb_2_p(row["Denominator:"]),
        #     "disaggregation_description": lb_2_p(row["Disaggregation(s) 1:"])+lb_2_p(row["Disaggregation(s) 2:"]),
        }
        de, c = models.DataElement.objects.update_or_create(
            name=name,
            **defaults
        )
        register(de)
        q_name = get_col(row,'H')
        if q_name:
            q, c = mallard_qr.Question.objects.update_or_create(
                name=get_col(row,'H').value.split('?')[0],
                question_text=get_col(row,'H').value,
                defaults = {
                    "collected_data_element":de,
                },
            )
            q.save()
            register(q)

        if vd and q_name:
            rd, c = mallard_qr.ResponseDomain.objects.update_or_create(
                question=q,
                value_domain=vd,
            )
        

def process_indicators(wb):
    sheet = wb.get_sheet_by_name('Indicators')
    for row in sheet.iter_rows(row_offset=3):
        if row[2].value is None:
            continue
        name = get_col(row,'C').value
        print name

        defn = get_col(row,'I')
        if defn:
            defn = get_col(row,'I').value or default_definition
        else:
            defn = default_definition
        defaults = {    
            "definition": defn,
        #     "definition": row["Purpose:"],
        #     "references": row["References:"],
             "references": get_col(row,'AB').value,
        #     "rationale": lb_2_p(row["Interpretation:"]),
            "short_name": get_col(row,'E').value,
            "numerator_description": get_col(row,'P').value or "",
            "denominator_description": get_col(row,'Q').value or "",
        #     "disaggregation_description": lb_2_p(row["Disaggregation(s) 1:"])+lb_2_p(row["Disaggregation(s) 2:"]),
        }
        ind, c = comet.Indicator.objects.get_or_create(
            name=name,
            **defaults
        )
        register(ind)

        des = [de.strip() for de in get_col(row,'F').value.split(';')]
        for de in des:
            de = models.DataElement.objects.filter(short_name=de).first()
            if de:
                print "  -",de
                ind.numerators.add(de)
        
        instument_name = get_col(row,'AH').value
        instrument = lo_models.Instrument.objects.filter(name=instument_name)
        if instrument.exists():
            instrument.first().indicators.add(ind)
        else:
            print "No instrument", instument_name, instrument
        
        for col in 'OPQRSTUVWX':
            frameworks = get_col(row,'A'+col)
            if frameworks:
                frameworks = frameworks.value.split(';')
            else:
                frameworks = []
            for framework in frameworks:
                framework = framework.strip()
                if framework == "":
                    continue 
                f = comet.Framework.objects.filter(short_name__icontains=framework).first()
                if f:
                    ind.frameworks.add(f)
                else:
                    print ind.name, framework

        for sl in slot_columns:
            sl,name,desc = sl
            value = get_col(row,sl)
            if value:
                value = value.value
                obj,created = Slot.objects.get_or_create(
                    type=slots[sl],
                    concept=ind,
                    value=jpger(value)
                )


def process_instruments(wb):
    sheet = wb.get_sheet_by_name('Instruments')
    for row in sheet.iter_rows(row_offset=1):
        if row[2].value is None:
            continue
        name = get_col(row,'A').value

        #print repr(get_col(row,'I').value)

        defaults = {
            "definition": lb_2_p(get_col(row,'I').value or default_definition, sep="\n"),
            "terms_of_use": get_col(row,'G').value or "",
            "where_to_get": get_col(row,'F').value or "",
            "references": get_col(row,'E').value or "",
            "limitations": get_col(row,'D').value or "",
            "short_name": get_col(row,'H').value,
            "population": get_col(row,'C').value,
        #     "disaggregation_description": lb_2_p(row["Disaggregation(s) 1:"])+lb_2_p(row["Disaggregation(s) 2:"]),
        }
        inst, c = lo_models.Instrument.objects.update_or_create(
            name=name,
            **defaults
        )
        register(inst)


def process_frameworks(wb):
    sheet = wb.get_sheet_by_name('Financial topics')
    for row in sheet.iter_rows(row_offset=1):
        if row[0].value is None:
            continue
        name = get_col(row,'A').value

        defaults = {
            "definition": "",
            "short_name": get_col(row,'B').value,
        }
        f, c = comet.Framework.objects.update_or_create(
            name=name,
            **defaults
        )
        register(f)

def register(thing):
    models.Status.objects.get_or_create(
        concept=thing,
        registrationAuthority=lo,
        registrationDate=datetime.date(2009, 4, 28),
        state=models.STATES.standard
    )


def process_indicator_slots():
    
    for sl in slot_columns:
        sl,name,desc = sl
        slot_type,created = SlotDefinition.objects.get_or_create(
            app_label='comet',
            concept_type='indicator',
            slot_name=name,
            cardinality = SlotDefinition.CARDINALITY.repeatable
        )
        slots[sl]=slot_type
    return slots


def lb_2_p(txt, sep="\n\n"):
    if sep in txt:
        return "<p>"+"</p><p>".join([l for l in txt.split(sep) if l != ""])+"</p>"
    else:
        return txt

import re
def jpger(txt):
    txt = re.sub(r"(https?://.*?jpg)", "<img src='\\1'>", txt)
    return txt

def get_col(row,col):
    for cell in row:
        if cell.value is not None:
            if cell.coordinate.startswith(col):
                return cell


if __name__ == "__main__":
    # slots = process_indicator_slots()
    process_indicator_slots()

    print("starting")
    wb = load_workbook(xls_file, read_only=True)

    process_frameworks(wb)
    process_instruments(wb)
    process_value_domains(wb)
    process_data_elements(wb)
    process_indicators(wb)
    

if False:
    reader = unicode_csv_reader(csvfile)
    for row in reader:
        name = row[u"Indicator Name:"].strip()
        if name == "":
            continue 
        defaults = {
            "workgroup": wg,
            "short_name": row["Short name:"],
            "definition": row["Purpose:"],
            "references": row["References:"],
            "rationale": lb_2_p(row["Interpretation:"]),
            "numerator_description": lb_2_p(row["Numerator:"]),
            "denominator_description": lb_2_p(row["Denominator:"]),
            "disaggregation_description": lb_2_p(row["Disaggregation(s) 1:"])+lb_2_p(row["Disaggregation(s) 2:"]),
        }
        indicator = comet.Indicator.objects.create(
            name=name,
            **defaults
        )

        models.Status.objects.get_or_create(
            concept=indicator,
            registrationAuthority=lo,
            registrationDate=datetime.date(2009, 4, 28),
            state=models.STATES.standard
        )

        for sl in slot_columns:
            sl,name,desc = sl
            value = row[sl]
            obj,created = Slot.objects.get_or_create(
                type=slots[sl],
                concept=indicator,
                value=jpger(value)
            )
        
        # set relevant populations
        # populations = row["Universe:"].split(';')
        # for pop in populations:
        #     pop = pop.strip()
        #     if pop == "":
        #         continue 
        #     pop, create = models.ObjectClass.objects.get_or_create(
        #         name = pop,
        #         defaults = {
        
        #             "workgroup": wg,
        #             "definition": "",
        #         }
        #     )
        #     models.Status.objects.get_or_create(
        #         concept=pop,
        #         registrationAuthority=lo,
        #         registrationDate=datetime.date(2009, 4, 28),
        #         state=models.STATES.standard
        #     )
        #     pop, c = lo_models.Population.objects.get_or_create(object_class=pop,indicator=indicator)
        
        
        # set outcomes
        outcomes = row["Indicator group:"].split(';')
        for outcome in outcomes:
            outcome = outcome.strip()
            if outcome == "":
                continue 
            outcome, created = comet.OutcomeArea.objects.get_or_create(
                name = outcome,
                defaults = {
                    "workgroup": wg,
                    "definition": "",
                }
            )
            if created:
                models.Status.objects.get_or_create(
                    concept=outcome,
                    registrationAuthority=lo,
                    registrationDate=datetime.date(2009, 4, 28),
                    state=models.STATES.standard
                )
            indicator.outcome_areas.add(outcome)

        # set outcomes
        frameworks = row["Framework:"].split(';')
        for framework in frameworks:
            framework = framework.strip()
            if framework == "":
                continue 
            framework, created = comet.Framework.objects.get_or_create(
                name = framework,
                defaults = {
                    "workgroup": wg,
                    "definition": "",
                }
            )
            if created:
                models.Status.objects.get_or_create(
                    concept=framework,
                    registrationAuthority=lo,
                    registrationDate=datetime.date(2009, 4, 28),
                    state=models.STATES.standard
                )
            framework.indicators.add(indicator)

        if created:
            models.Status.objects.get_or_create(
                concept=framework,
                registrationAuthority=lo,
                registrationDate=datetime.date(2009, 4, 28),
                state=models.STATES.standard
            )
        framework.indicators.add(indicator)
        
if False:  #with open(vd_file,'rb') as csvfile:
    #vds = csvfile.read().split('\n\t\t\t\n')
    val_dom = None
    indicator = None
    n = 0
    for i,row in enumerate(unicode_csv_reader(csvfile,reader=csv.reader)):
        if val_dom is None:
            ind_code = row[0]
            vd_name = row[1]
            meaning = row[2]
            value = row[3]
            dt = models.DataType.objects.get(name='Number')
            val_dom, created = models.ValueDomain.objects.get_or_create(
                    name = vd_name,
                    defaults = {
                        "workgroup": wg,
                        "definition": "Imported from the Indicator Reference Sheet",
                        "data_type": dt,
                    }
                )
            models.Status.objects.get_or_create(
                concept=val_dom,
                registrationAuthority=lo,
                registrationDate=datetime.date(2009, 4, 28),
                state=models.STATES.standard
            )
            indicator = comet.Indicator.objects.filter(slots__type=slots['Indicator Code:'],slots__value=ind_code).first()
            if indicator:
                indicator.valueDomain = val_dom
                indicator.save()
            else:
                print "No indicator - ", ind_code
            models.PermissibleValue.objects.create(valueDomain=val_dom,value=value,meaning=meaning,order=n)
            n += 1
        elif row == ['','','','']:
            if n == 0:
                val_dom.permssible_values.delete()
            val_dom = None
            indicator = None
            n = 0
        else:
            meaning = row[2]
            value = row[3]
            models.PermissibleValue.objects.create(valueDomain=val_dom,value=value,meaning=meaning,order=n)
            n += 1
